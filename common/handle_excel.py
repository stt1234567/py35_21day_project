'''
封装的需求：
    1、数据读取:封装一个可以读取任意excel文件的方法，可以指定读取的表单
    2、数据写入:

'''
import openpyxl
class HandleExcel:
    def __init__(self,filename,sheetname):
        """
        :param filename:写入文件地址
        :param sheetname: 写入表单sheet名
        """
        self.filename=filename
        self.sheetname=sheetname
    def read_data(self):
        #------------读取excel数据----------
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        li = []
        title1 = [i.value for i in res[0]]
        for item in res[1:]:
            title2 = [i.value for i in item]
            dic = dict(zip(title1, title2))
            li.append(dic)
        return li
    def write_data(self,row,column,value):
        """

        :param row: 写入行
        :param column:写入列
        :param value:写入值
        :return:
        """
        #------------写入excel数据----------
        #载入工作簿对象
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        #写入数据
        sh.cell(row=row,column=column,value=value)
        workbook.save(self.filename)


# if __name__ == '__main__':
#     excel=HandleExcel(r'D:\work\zx5jkzdh\jichu\py35_17day_project\datas\test002.xlsx', 'Sheet1')
#     res=excel.read_data()
#     print(res)